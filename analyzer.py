"""
NUCLEUS ANALYZER - Your original working code, using config
"""

from collections import defaultdict
from vedo import Volume
import numpy as np
import imageio
from scipy.ndimage import zoom, gaussian_filter
import trimesh
import pandas as pd
import os 
import math
import warnings
import gc
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import Workbook

# Import from config and detector
from config import *
from detector import detect_group_from_filename, get_groups_ordered

warnings.filterwarnings('ignore')


class NucleusProcessor:
    def __init__(self):
        self.original_spacing = ORIGINAL_SPACING
        self.target_spacing = TARGET_SPACING
        self.curv_r = CURVATURE_RADIUS
        self.bins = CURVATURE_BINS
        self.z_exclusion = Z_EXCLUSION
        
        self.zoom_factors = (
            ORIGINAL_SPACING[0] / TARGET_SPACING,
            ORIGINAL_SPACING[1] / TARGET_SPACING,
            ORIGINAL_SPACING[2] / TARGET_SPACING
        )
        self.sigma = 0.8 / TARGET_SPACING
        self.voxel_volume_um3 = ORIGINAL_SPACING[0] * ORIGINAL_SPACING[1] * ORIGINAL_SPACING[2]
        self.voxel_volume_pL = self.voxel_volume_um3 * 0.001
    
    def crop_to_bbox(self, mask, margin=8):
        coords = np.argwhere(mask)
        if len(coords) == 0:
            return None
        
        xmin, ymin, zmin = coords.min(axis=0)
        xmax, ymax, zmax = coords.max(axis=0) + 1
        
        xmin = max(xmin - margin, 0)
        ymin = max(ymin - margin, 0)
        zmin = max(zmin - margin, 0)
        
        xmax = min(xmax + margin, mask.shape[0])
        ymax = min(ymax + margin, mask.shape[1])
        zmax = min(zmax + margin, mask.shape[2])
        
        return mask[xmin:xmax, ymin:ymax, zmin:zmax]
    
    def process_single_nucleus(self, mask_data):
        """Process one nucleus - returns data for Excel"""
        mask, label, file_path, embryo_id, nucleus_id, group, filename = mask_data
        
        try:
            # 1. Resampling
            data_isotropic = zoom(mask.astype(np.float32), zoom=self.zoom_factors, order=1)
            
            # 2. Physical smoothing
            data_isotropic = gaussian_filter(data_isotropic, 
                                            sigma=(self.sigma, self.sigma, self.sigma))
            
            # 3. Volume calculation
            voxels_nb = np.sum(mask > 0)
            nucleus_volume_pL = voxels_nb * self.voxel_volume_pL
            
            # 4. Mesh creation
            vol = Volume(data_isotropic, spacing=(self.target_spacing,)*3)
            surf = vol.isosurface(0.5)
            
            if surf.npoints < 100:
                return None
            
            # Your smoothing
            surf.clean()
            surf.decimate(0.5, preserve_volume=True)
            surf.smooth(niter=20, pass_band=0.15)
            
            # 5. Convert to trimesh
            vertices = np.ascontiguousarray(surf.vertices, dtype=np.float32)
            faces = np.ascontiguousarray(surf.cells, dtype=np.int32)
            surf_trimesh = trimesh.Trimesh(vertices=vertices, faces=faces)
            
            # Calculate per-vertex Voronoi areas
            tri_areas = surf_trimesh.area_faces
            tri_vertices = surf_trimesh.faces
            vertex_areas_all = np.zeros(len(surf_trimesh.vertices))
            area_contribs = tri_areas / 3.0
            
            for tri_idx, (v1, v2, v3) in enumerate(tri_vertices):
                contrib = area_contribs[tri_idx]
                vertex_areas_all[v1] += contrib
                vertex_areas_all[v2] += contrib
                vertex_areas_all[v3] += contrib
                
            # 6. Curvature
            curvatures = trimesh.curvature.discrete_mean_curvature_measure(
                surf_trimesh, surf_trimesh.vertices, self.curv_r
            )
            
            # 7. Z exclusion
            z_coords = surf_trimesh.vertices[:, 2]
            z_min, z_max = z_coords.min(), z_coords.max()
            z_range = z_max - z_min
            z_lower = z_min + self.z_exclusion * z_range
            z_upper = z_max - self.z_exclusion * z_range
            
            central_mask = (z_coords >= z_lower) & (z_coords <= z_upper)
            curv = np.array(curvatures, dtype=np.float32)
            curv[~central_mask] = np.nan
            curv_central = curv[~np.isnan(curv)]
            areas_central = vertex_areas_all[central_mask]
            
            if curv_central.size < 100:
                return None
            
            # 8. Area-weighted histogram
            counts, bin_edges = np.histogram(
                curv_central, 
                bins=self.bins,
                weights=areas_central
            )
            bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
            
            # 9. Sphere comparison
            radius_um = (3 * nucleus_volume_pL * 1000 / (4 * np.pi)) ** (1/3)
            sphere = trimesh.creation.icosphere(subdivisions=3, radius=radius_um)
            sphere_curv = np.mean(trimesh.curvature.discrete_mean_curvature_measure(
                sphere, sphere.vertices, self.curv_r
            ))
            
            # 10. Surface percentages
            counts_pos = np.sum(counts[bin_centers >= 0])
            counts_neg = np.sum(counts[bin_centers < 0])
            total = counts_pos + counts_neg
            
            # 11. Weighted mean absolute curvature
            if len(curv_central) > 0:
                weighted_mean_abs_curv = np.average(
                    np.abs(curv_central), 
                    weights=areas_central
                )
            else:
                weighted_mean_abs_curv = 0
        
            # Create unique ID
            nucleus_name = f"{group}_Embryo{embryo_id:02d}_Nucleus{nucleus_id:02d}"
            
            # Clean up
            del data_isotropic, surf, vertices, faces, sphere
            if 'surf_trimesh' in locals():
                del surf_trimesh
            
            return {
                'nucleus_name': nucleus_name,
                'group': group,
                'embryo_id': embryo_id,
                'nucleus_id': nucleus_id,
                'filename': filename,
                'label_in_image': label,
                'curvature_bin_centers': bin_centers,
                'curvature_counts': counts,
                'volume_pL': nucleus_volume_pL,
                'weighted_mean_abs_curv': weighted_mean_abs_curv,
                'sphere_curvature': sphere_curv,
                'theoretical_sphere_curvature': (2 * math.pi * (radius_um - (radius_um**2 - self.curv_r**2)**(1/2))),
                'positive_surface_percent': counts_pos / total * 100 if total > 0 else 0,
                'negative_surface_percent': counts_neg / total * 100 if total > 0 else 0,
                'voxel_count': voxels_nb,
                'n_valid_points': len(curv_central),
                'n_total_points': len(curv),
                'curvature_radius_um': self.curv_r
            }
            
        except Exception as e:
            print(f"    Error in nucleus {label}: {str(e)[:100]}...")
            return None
        finally:
            gc.collect()


class ExcelStageWriter:
    def __init__(self, output_excel):
        self.output_excel = output_excel
        self.stage_data = defaultdict(lambda: {
            'curvature_data': [],
            'volume_data': [],
            'surface_data': []
        })
        self.processed_files = set()
        
    def add_nucleus_data(self, nucleus_data, filename):
        if filename in self.processed_files:
            print(f"  ⚠ File {filename} already processed - skipping")
            return False
        
        group = nucleus_data['group']
        
        self.stage_data[group]['curvature_data'].append({
            'nucleus_name': nucleus_data['nucleus_name'],
            'bin_centers': nucleus_data['curvature_bin_centers'],
            'counts': nucleus_data['curvature_counts']
        })
        
        self.stage_data[group]['volume_data'].append({
            'nucleus_name': nucleus_data['nucleus_name'],
            'embryo_id': nucleus_data['embryo_id'],
            'nucleus_id': nucleus_data['nucleus_id'],
            'filename': nucleus_data['filename'],
            'volume_pL': nucleus_data['volume_pL'],
            'weighted_mean_abs_curv': nucleus_data['weighted_mean_abs_curv'],
            'sphere_curvature': nucleus_data['sphere_curvature'],
            'theoretical_sphere_curvature': nucleus_data['theoretical_sphere_curvature'],
            'voxel_count': nucleus_data['voxel_count'],
            'curvature_radius_um': nucleus_data['curvature_radius_um']
        })
        
        self.stage_data[group]['surface_data'].append({
            'nucleus_name': nucleus_data['nucleus_name'],
            'embryo_id': nucleus_data['embryo_id'],
            'nucleus_id': nucleus_data['nucleus_id'],
            'positive_surface_percent': nucleus_data['positive_surface_percent'],
            'negative_surface_percent': nucleus_data['negative_surface_percent'],
            'n_valid_points': nucleus_data['n_valid_points'],
            'n_total_points': nucleus_data['n_total_points']
        })
        
        return True
    
    def mark_file_processed(self, filename):
        self.processed_files.add(filename)
    
    def save_all_sheets(self):
        print("\nCreating final Excel file...")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        groups_order = get_groups_ordered()
        
        for group in groups_order:
            if group not in self.stage_data or not self.stage_data[group]['curvature_data']:
                continue
            
            print(f"  Creating sheets for {group}...")
            self._create_curvature_sheet(wb, group)
            self._create_volume_sheet(wb, group)
            self._create_surface_sheet(wb, group)
        
        self._create_summary_sheet(wb)
        
        wb.save(self.output_excel)
        print(f"\n✓ Excel file saved: {self.output_excel}")
        
        total_nuclei = sum(len(data['curvature_data']) for data in self.stage_data.values())
        print(f"✓ Total nuclei processed: {total_nuclei}")
    
    def _create_curvature_sheet(self, wb, group):
        ws = wb.create_sheet(f"Curvature_data_{group}")
        curvature_data = self.stage_data[group]['curvature_data']
        
        col = 1
        for nucleus in curvature_data:
            ws.cell(row=1, column=col, value=f"{nucleus['nucleus_name']}_Curvature")
            ws.cell(row=1, column=col + 1, value=f"{nucleus['nucleus_name']}_Counts")
            col += 2
        
        for row_idx in range(CURVATURE_BINS):
            col = 1
            for nucleus in curvature_data:
                if row_idx < len(nucleus['bin_centers']):
                    ws.cell(row=row_idx + 2, column=col, value=nucleus['bin_centers'][row_idx])
                    ws.cell(row=row_idx + 2, column=col + 1, value=nucleus['counts'][row_idx])
                col += 2
        
        print(f"    ✓ Curvature_data_{group}: {len(curvature_data)} nuclei")
    
    def _create_volume_sheet(self, wb, group):
        ws = wb.create_sheet(f"Volume_data_{group}")
        volume_data = self.stage_data[group]['volume_data']
        
        headers = ["Nucleus Name", "Embryo ID", "Nucleus ID", "Filename",
                   "Volume (pL)", "Weighted Mean |Curvature| (µm)",
                   "Sphere Curvature (µm)", "Theoretical Sphere Curvature (µm)",
                   "Voxel Count", "Curvature Radius (µm)"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, nucleus in enumerate(volume_data, 2):
            ws.cell(row=row_idx, column=1, value=nucleus['nucleus_name'])
            ws.cell(row=row_idx, column=2, value=nucleus['embryo_id'])
            ws.cell(row=row_idx, column=3, value=nucleus['nucleus_id'])
            ws.cell(row=row_idx, column=4, value=nucleus['filename'])
            ws.cell(row=row_idx, column=5, value=nucleus['volume_pL'])
            ws.cell(row=row_idx, column=6, value=nucleus['weighted_mean_abs_curv'])
            ws.cell(row=row_idx, column=7, value=nucleus['sphere_curvature'])
            ws.cell(row=row_idx, column=8, value=nucleus['theoretical_sphere_curvature'])
            ws.cell(row=row_idx, column=9, value=nucleus['voxel_count'])
            ws.cell(row=row_idx, column=10, value=nucleus['curvature_radius_um'])
        
        print(f"    ✓ Volume_data_{group}: {len(volume_data)} nuclei")
    
    def _create_surface_sheet(self, wb, group):
        ws = wb.create_sheet(f"Surface_data_{group}")
        surface_data = self.stage_data[group]['surface_data']
        
        headers = ["Nucleus Name", "Embryo ID", "Nucleus ID",
                   "Positive Surface (%)", "Negative Surface (%)",
                   "N Valid Points", "N Total Points"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, nucleus in enumerate(surface_data, 2):
            ws.cell(row=row_idx, column=1, value=nucleus['nucleus_name'])
            ws.cell(row=row_idx, column=2, value=nucleus['embryo_id'])
            ws.cell(row=row_idx, column=3, value=nucleus['nucleus_id'])
            ws.cell(row=row_idx, column=4, value=nucleus['positive_surface_percent'])
            ws.cell(row=row_idx, column=5, value=nucleus['negative_surface_percent'])
            ws.cell(row=row_idx, column=6, value=nucleus['n_valid_points'])
            ws.cell(row=row_idx, column=7, value=nucleus['n_total_points'])
        
        print(f"    ✓ Surface_data_{group}: {len(surface_data)} nuclei")
    
    def _create_summary_sheet(self, wb):
        ws = wb.create_sheet("Summary")
        
        headers = ["Group", "N Nuclei", "Mean Volume (pL)", "Std Volume",
                   "Mean |Curvature| (µm)", "Std |Curvature|",
                   "Mean Positive Surface (%)", "Std Positive Surface (%)",
                   "Volume Range", "Curvature Range"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        row_idx = 2
        for group in get_groups_ordered():
            if group not in self.stage_data:
                continue
            
            volume_data = self.stage_data[group]['volume_data']
            surface_data = self.stage_data[group]['surface_data']
            
            if not volume_data:
                continue
            
            volumes = [n['volume_pL'] for n in volume_data]
            curvatures = [n['weighted_mean_abs_curv'] for n in volume_data]
            pos_surfaces = [n['positive_surface_percent'] for n in surface_data]
            
            ws.cell(row=row_idx, column=1, value=group)
            ws.cell(row=row_idx, column=2, value=len(volume_data))
            ws.cell(row=row_idx, column=3, value=np.mean(volumes))
            ws.cell(row=row_idx, column=4, value=np.std(volumes))
            ws.cell(row=row_idx, column=5, value=np.mean(curvatures))
            ws.cell(row=row_idx, column=6, value=np.std(curvatures))
            ws.cell(row=row_idx, column=7, value=np.mean(pos_surfaces))
            ws.cell(row=row_idx, column=8, value=np.std(pos_surfaces))
            ws.cell(row=row_idx, column=9, value=f"{min(volumes):.2f}-{max(volumes):.2f}")
            ws.cell(row=row_idx, column=10, value=f"{min(curvatures):.3f}-{max(curvatures):.3f}")
            
            row_idx += 1
        
        print(f"    ✓ Summary sheet created")


def process_embryo_parallel(processor, excel_writer, file_path_img, file_counter, max_workers=2):
    """Process one embryo with parallel processing"""
    filename = os.path.basename(file_path_img)
    group = detect_group_from_filename(filename)
    
    print(f"  File: {filename}")
    print(f"  Detected group: {group}")
    
    if filename in excel_writer.processed_files:
        print(f"  ⚠ File already processed - skipping")
        return 0
    
    print(f"  Processing with {max_workers} workers...")
    
    try:
        data_0 = imageio.volread(file_path_img)
        data_0 = np.transpose(data_0, (2, 1, 0)).astype(np.uint8)
        
        unique_labels = np.unique(data_0)
        unique_labels = unique_labels[unique_labels != 0]
        
        if len(unique_labels) == 0:
            print(f"  ⚠ No nuclei found")
            return 0
        
        mask_data_list = []
        nucleus_counter = 0
        
        for label in unique_labels:
            mask = (data_0 == label).astype(np.uint8)
            cropped = processor.crop_to_bbox(mask, margin=8)
            
            if cropped is not None and np.sum(cropped) > 100:
                mask_data_list.append((cropped, label, file_path_img, 
                                      file_counter, nucleus_counter + 1, 
                                      group, filename))
                nucleus_counter += 1
        
        if not mask_data_list:
            print(f"  ⚠ No valid nuclei after cropping")
            return 0
        
        print(f"  Found {len(mask_data_list)} valid nuclei")
        
        processed_count = 0
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(processor.process_single_nucleus, md): md 
                      for md in mask_data_list}
            
            completed = 0
            for future in as_completed(futures):
                try:
                    result = future.result(timeout=300)
                    if result:
                        if excel_writer.add_nucleus_data(result, filename):
                            processed_count += 1
                    
                    completed += 1
                    if completed % 5 == 0:
                        print(f"    Completed {completed}/{len(mask_data_list)} nuclei")
                        
                except Exception as e:
                    print(f"    Task failed: {str(e)[:100]}...")
                    completed += 1
        
        if processed_count > 0:
            excel_writer.mark_file_processed(filename)
        
        del data_0
        gc.collect()
        
        return processed_count
        
    except Exception as e:
        print(f"  ✗ Error processing embryo: {e}")
        import traceback
        traceback.print_exc()
        return 0


def main():
    print("\n" + "="*70)
    print(EXPERIMENT_NAME)
    print("="*70)
    print(f"Results will be saved to: {OUTPUT_EXCEL}")
    print("="*70 + "\n")
    
    processor = NucleusProcessor()
    excel_writer = ExcelStageWriter(OUTPUT_EXCEL)
    
    files = [os.path.join(INPUT_FOLDER, f) for f in os.listdir(INPUT_FOLDER) 
             if f.lower().endswith(('.tif', '.tiff'))]
    files.sort()
    
    print(f"Found {len(files)} image files")
    
    print("\nGroup detection verification:")
    group_counts = defaultdict(int)
    for file_path_img in files:
        filename = os.path.basename(file_path_img)
        group = detect_group_from_filename(filename)
        group_counts[group] += 1
        print(f"  {filename[:50]}... → {group}")
    
    print("\nGroup summary:")
    for group in get_groups_ordered():
        if group_counts[group] > 0:
            print(f"  {group}: {group_counts[group]} files")
    
    total_processed = 0
    file_processed = 0
    
    for file_idx, file_path_img in enumerate(files, 1):
        filename = os.path.basename(file_path_img)
        
        print(f"\n{'='*60}")
        print(f"File {file_idx}/{len(files)}: {filename}")
        print('='*60)
        
        try:
            processed = process_embryo_parallel(
                processor, excel_writer, file_path_img, file_idx, MAX_WORKERS
            )
            
            if processed > 0:
                total_processed += processed
                file_processed += 1
                print(f"  ✓ Processed {processed} nuclei from this file")
            else:
                print("  ⚠ No nuclei processed from this file")
            
            gc.collect()
            
        except Exception as e:
            print(f"  ✗ Failed to process file: {e}")
            continue
    
    print(f"\n{'='*60}")
    print("Creating final Excel file...")
    print('='*60)
    
    excel_writer.save_all_sheets()
    
    print(f"\n{'='*70}")
    print(f"PROCESSING COMPLETE")
    print(f"{'='*70}")
    print(f"Total files processed: {file_processed}/{len(files)}")
    print(f"Total nuclei processed: {total_processed}")
    print(f"\nExcel file: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()